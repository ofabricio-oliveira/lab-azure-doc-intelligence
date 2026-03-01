"""
Unit tests for app.utils.exporter (CSV and XLSX generation).
"""

import csv
import io
import pytest
from openpyxl import load_workbook

from app.models import AnalysisResult, LineItem
from app.utils.exporter import generate_csv, generate_xlsx, COLUMNS, ITEM_COLUMNS


@pytest.fixture
def sample_result():
    """A typical AnalysisResult for testing exports."""
    return AnalysisResult(
        analysis_id="exp001",
        document_type="receipt",
        vendor_name="Contoso Cafe",
        document_date="2025-01-20",
        currency="USD",
        subtotal=14.00,
        tax=1.50,
        total=15.50,
        document_id=None,
        items=[
            LineItem(description="Cappuccino", quantity=2, unit_price=5.00, amount=10.00, confidence=0.93),
            LineItem(description="Muffin", quantity=1, unit_price=4.00, amount=4.00, confidence=0.91),
        ],
        confidence_score=0.92,
        needs_review=False,
        source_filename="receipt.pdf",
    )


@pytest.fixture
def empty_items_result():
    """Result with no line items."""
    return AnalysisResult(
        analysis_id="exp002",
        document_type="invoice",
        vendor_name="ACME Corp",
        document_date="2025-06-15",
        total=1200.00,
        needs_review=False,
        source_filename="invoice.pdf",
    )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

class TestGenerateCSV:
    def test_csv_has_header(self, sample_result):
        csv_text = generate_csv(sample_result)
        reader = csv.reader(io.StringIO(csv_text))
        header = next(reader)
        for col in COLUMNS + ITEM_COLUMNS:
            assert col in header

    def test_csv_row_count_matches_items(self, sample_result):
        csv_text = generate_csv(sample_result)
        reader = csv.DictReader(io.StringIO(csv_text))
        rows = list(reader)
        assert len(rows) == 2  # 2 items

    def test_csv_values(self, sample_result):
        csv_text = generate_csv(sample_result)
        reader = csv.DictReader(io.StringIO(csv_text))
        row = next(reader)
        assert row["vendor_name"] == "Contoso Cafe"
        assert row["total"] == "15.5"
        assert row["needs_review"] == "false"
        assert row["item_description"] == "Cappuccino"

    def test_csv_no_items_single_row(self, empty_items_result):
        csv_text = generate_csv(empty_items_result)
        reader = csv.DictReader(io.StringIO(csv_text))
        rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["vendor_name"] == "ACME Corp"
        assert rows[0]["item_description"] == ""


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

class TestGenerateXLSX:
    def test_xlsx_is_valid(self, sample_result):
        xlsx_bytes = generate_xlsx(sample_result)
        assert len(xlsx_bytes) > 0
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws is not None

    def test_xlsx_header(self, sample_result):
        xlsx_bytes = generate_xlsx(sample_result)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        for col in COLUMNS + ITEM_COLUMNS:
            assert col in header

    def test_xlsx_row_count(self, sample_result):
        xlsx_bytes = generate_xlsx(sample_result)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # 1 header + 2 data rows
        assert ws.max_row == 3

    def test_xlsx_no_items(self, empty_items_result):
        xlsx_bytes = generate_xlsx(empty_items_result)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.max_row == 2  # 1 header + 1 data row
